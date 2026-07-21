"""
Faz login em http://172.30.0.38:8080/changelog/ e mantém uma cópia local
atualizada do changelog de políticas (Crefaz).

O documento servido pelo site já acumula, num único HTML, todo o histórico de
versões desde o início (V220913 em diante) até a versão mais recente e a seção
"Não Publicado" — não há paginação/versionamento separado a buscar. Este
script portanto:

1. Loga no site e busca o HTML completo do changelog (via iframe de /changelog/).
2. Compara com o hash e a lista de versões salvos da última execução.
3. Se algo mudou, salva um snapshot datado, atualiza a cópia "completo" e o
   índice em Excel, e registra no log quais versões são novas.
4. Se nada mudou, apenas registra o "ping" no log e sai.

Uso:
    python atualizar_changelog_politicas.py

Credenciais em ".env" (mesma pasta, ver ".env.example").
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter

import gerar_changelog_completo_excel as gerador_completo
import gerar_changelog_tipo_mudanca_excel as gerador_tipo_mudanca


def encontrar_raiz_projeto(inicio: Path) -> Path:
    for pasta in [inicio.resolve(), *inicio.resolve().parents]:
        if (pasta / "CLAUDE.md").exists():
            return pasta
    raise RuntimeError("CLAUDE.md não encontrado subindo a árvore de diretórios")


SCRIPT_DIR = Path(__file__).resolve().parent
RAIZ = encontrar_raiz_projeto(SCRIPT_DIR)

SAIDA_DIR = RAIZ / "changelog politicas"
SNAPSHOTS_DIR = SAIDA_DIR / "snapshots"
ESTADO_PATH = SAIDA_DIR / "estado.json"
COMPLETO_PATH = SAIDA_DIR / "changelog_completo.html"
INDICE_XLSX_PATH = SAIDA_DIR / "index_versoes.xlsx"
LOG_PATH = SAIDA_DIR / "log_execucao.log"

load_dotenv(SCRIPT_DIR / ".env")

BASE_URL = os.environ["CHANGELOG_BASE_URL"].rstrip("/")
USUARIO = os.environ["CHANGELOG_USUARIO"]
SENHA = os.environ["CHANGELOG_SENHA"]

SAIDA_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("changelog_politicas")


def login(session: requests.Session) -> None:
    resp = session.get(f"{BASE_URL}/", timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    token_input = soup.find("input", {"name": "csrfmiddlewaretoken"})
    if token_input is None:
        raise RuntimeError("Formulário de login não encontrado em / (layout do site mudou?)")

    session.post(
        f"{BASE_URL}/",
        data={
            "csrfmiddlewaretoken": token_input["value"],
            "usuario": USUARIO,
            "password": SENHA,
        },
        headers={"Referer": f"{BASE_URL}/"},
        timeout=15,
        allow_redirects=False,
    )
    if "sessionid" not in session.cookies.get_dict():
        raise RuntimeError("Login falhou: usuário/senha inválidos ou layout do site mudou")


def buscar_documento_changelog(session: requests.Session) -> tuple[str, str]:
    resp = session.get(f"{BASE_URL}/changelog/", timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    iframe = soup.find("iframe")
    if iframe is None or not iframe.get("src"):
        raise RuntimeError("iframe do changelog não encontrado em /changelog/ (layout do site mudou?)")

    iframe_src = iframe["src"].split("#")[0]
    iframe_url = iframe_src if iframe_src.startswith("http") else f"{BASE_URL}{iframe_src}"

    resp_doc = session.get(iframe_url, timeout=15)
    resp_doc.raise_for_status()
    # O servidor não declara charset no header Content-Type; o documento é UTF-8
    # (confirmado pelo <meta charset="UTF-8"> da página), então o requests
    # cairia no fallback ISO-8859-1 e corromperia acentos sem isto.
    resp_doc.encoding = "utf-8"
    return iframe_url, resp_doc.text


def parse_versoes(html: str) -> list[dict]:
    """Extrai a lista de versões a partir dos links do Sumário (ordem: mais nova -> mais antiga)."""
    soup = BeautifulSoup(html, "html.parser")
    versoes = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not href.startswith("#"):
            continue
        ancora = href[1:]
        if ancora != "NaoPublicado" and not re.fullmatch(r"V\d+", ancora):
            continue
        texto = a.get_text(strip=True)
        data = texto.rsplit(" - ", 1)[-1].strip() if " - " in texto else None
        versoes.append({"id": ancora, "titulo": texto, "data": data})
    return versoes


def carregar_estado() -> dict:
    if ESTADO_PATH.exists():
        return json.loads(ESTADO_PATH.read_text(encoding="utf-8"))
    return {"hash_documento": None, "versoes_conhecidas": []}


def salvar_estado(estado: dict) -> None:
    ESTADO_PATH.write_text(json.dumps(estado, ensure_ascii=False, indent=2), encoding="utf-8")


def atualizar_indice_excel(versoes: list[dict], novas_ids: set[str]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Versoes"
    ws.append(["Versao", "Data", "Novo nesta execucao"])
    for v in reversed(versoes):  # mais antiga -> mais nova, como um changelog de verdade
        ws.append([v["titulo"], v["data"] or "", "SIM" if v["id"] in novas_ids else ""])
    for col_idx, largura in enumerate([30, 15, 20], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    wb.save(INDICE_XLSX_PATH)


def main() -> None:
    SNAPSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    estado = carregar_estado()
    primeira_execucao = estado.get("hash_documento") is None

    session = requests.Session()
    login(session)
    iframe_url, html = buscar_documento_changelog(session)

    hash_atual = hashlib.sha256(html.encode("utf-8")).hexdigest()
    versoes = parse_versoes(html)
    ids_atuais = {v["id"] for v in versoes}
    ids_conhecidas = set(estado.get("versoes_conhecidas", []))
    novas_ids = set() if primeira_execucao else (ids_atuais - ids_conhecidas)

    mudou_conteudo = hash_atual != estado.get("hash_documento")
    agora = datetime.now()

    if not mudou_conteudo:
        log.info(
            "Nada novo no changelog de políticas (ping ok, %d versões conhecidas, sem alteração de conteúdo).",
            len(ids_atuais),
        )
        estado["ultima_verificacao"] = agora.isoformat()
        salvar_estado(estado)
        return

    snapshot_path = SNAPSHOTS_DIR / f"changelog_{agora.strftime('%Y-%m-%d_%H%M%S')}.html"
    snapshot_path.write_text(html, encoding="utf-8")
    COMPLETO_PATH.write_text(html, encoding="utf-8")
    atualizar_indice_excel(versoes, novas_ids)
    gerador_completo.main()
    gerador_tipo_mudanca.main()

    if primeira_execucao:
        mais_antiga = versoes[-1]["titulo"] if versoes else "?"
        mais_nova = versoes[0]["titulo"] if versoes else "?"
        log.info(
            "Carga inicial do changelog de políticas: %d versões capturadas (de %s até %s).",
            len(ids_atuais), mais_antiga, mais_nova,
        )
    elif novas_ids:
        titulos_novos = [v["titulo"] for v in versoes if v["id"] in novas_ids]
        log.info(
            "Changelog de políticas atualizado: %d versão(ões) nova(s): %s",
            len(novas_ids), ", ".join(titulos_novos),
        )
    else:
        log.info(
            "Changelog de políticas atualizado: conteúdo mudou mas nenhuma versão nova detectada "
            "(edição em versão existente) — revisar %s",
            snapshot_path.name,
        )

    estado.update(
        {
            "hash_documento": hash_atual,
            "versoes_conhecidas": sorted(ids_atuais),
            "ultima_verificacao": agora.isoformat(),
            "ultima_atualizacao": agora.isoformat(),
            "iframe_url_atual": iframe_url,
        }
    )
    salvar_estado(estado)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        log.exception("Falha ao verificar changelog de políticas")
        raise
