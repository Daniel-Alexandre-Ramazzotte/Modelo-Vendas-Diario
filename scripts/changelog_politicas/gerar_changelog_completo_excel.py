"""
Gera "changelog politicas/changelog_politicas_completo.xlsx" a partir do HTML
completo do changelog de políticas ("changelog politicas/changelog_completo.html",
mantido por atualizar_changelog_politicas.py).

Mesmo "tipo" de planilha que changelog_energia.xlsx (colunas Versão, Data, Produto,
Política, Nome da Regra, CIA, Abrangência, Tipo de Mudança, Texto da Mudança), mas
cobrindo TODOS os produtos, não só Energia.

Diferenças conscientes em relação ao changelog_energia.xlsx original (decisão do
usuário em 2026-07-20):
  - CIA / Abrangência só são preenchidas para produtos de Energia (CP Energia, CDC
    Energia, ECP), onde a distribuidora aparece no texto da regra. Para os demais
    produtos esse conceito não existe, então ficam em branco.
  - Tipo de Mudança fica em branco para todos os produtos: no changelog_energia.xlsx
    essa coluna parece resultado de classificação manual sobre o texto, sem regra
    mecânica recuperável do HTML — preenchê-la por conta própria arriscaria inventar
    uma categoria errada num documento de política de crédito.
  - Seções "Alteração(ões) Crivo" são excluídas, assim como no changelog_energia.xlsx
    original (são mudanças de motor/infraestrutura, não de política).

Uso:
    python gerar_changelog_completo_excel.py
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Esquema de cores replicado do changelog_energia.xlsx original (mesmas cores para
# as categorias já existentes lá; cores novas, no mesmo estilo pastel, para as
# categorias que só existem no changelog "de tudo").
COR_CABECALHO = PatternFill("solid", fgColor="1F3864")
FONTE_CABECALHO = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
COR_FAIXA_PAR = PatternFill("solid", fgColor="DCE6F1")
COR_FAIXA_IMPAR = PatternFill("solid", fgColor="FFFFFF")
_LADO_FINO = Side(style="thin", color="BFBFBF")
BORDA_FINA = Border(left=_LADO_FINO, right=_LADO_FINO, top=_LADO_FINO, bottom=_LADO_FINO)

PRODUTO_CORES = {
    "CP Energia": "DDEEFF",
    "CDC Energia": "E8F5E9",
    "CP Energia + CDC Energia": "EDE7F6",
    "Energia e CDC Energia": "D1C4E9",
    "ECP": "FFF3CD",
    "CP Auto": "FFE0B2",
    "CP Boleto": "D7CCC8",
    "CP Refin": "F8BBD0",
    "Débito": "B2EBF2",
    "Financiamento Crefaz": "C5CAE9",
    "Saúde": "FFCCBC",
    "Todos": "E0E0E0",
    "Renda Presumida": "DCEDC8",
    "CDC": "FFECB3",
}

# Prioridade quando há mais de uma CIA na mesma linha (mesmo critério observado no
# changelog_energia.xlsx: a cor não é uma mistura, é a da CIA de maior prioridade).
CIA_PRIORIDADE = ["CPFL", "ENEL SP", "ENEL RJ", "ENEL CE", "ENEL", "NEO", "COSERN", "RGE"]
CIA_CORES = {
    "CPFL": "C6EFCE",
    "ENEL SP": "FFEB9C",
    "ENEL RJ": "FFEB9C",
    "ENEL CE": "FFEB9C",
    "ENEL": "FFEB9C",
    "NEO": "FFC7CE",
    "COSERN": "FCE4D6",
    "RGE": "D9D2E9",
    "Todas": "EDEDED",
}

TIPO_CORES = {
    "Bloqueio/Negativa": "FFC7CE",
    "Outros": "F2F2F2",
    "Novo produto/regra": "FCE4D6",
    "API/Integração": "E2EFDA",
    "Validação": "BDD7EE",
    "Aprovação automática": "C6EFCE",
    "Fluxo": "EBF3FB",
    "Limite de crédito": "FFEB9C",
    "Variável": "D9D9D9",
}


def _fill(cor_hex: str) -> PatternFill:
    return PatternFill("solid", fgColor=cor_hex)


def _cor_cia(cia: str) -> str | None:
    if not cia:
        return None
    presentes = {c.strip() for c in cia.split(", ")}
    if "Todas" in presentes:
        return CIA_CORES["Todas"]
    for prioridade in CIA_PRIORIDADE:
        if prioridade in presentes:
            return CIA_CORES[prioridade]
    return None


def encontrar_raiz_projeto(inicio: Path) -> Path:
    for pasta in [inicio.resolve(), *inicio.resolve().parents]:
        if (pasta / "CLAUDE.md").exists():
            return pasta
    raise RuntimeError("CLAUDE.md não encontrado subindo a árvore de diretórios")


RAIZ = encontrar_raiz_projeto(Path(__file__).resolve().parent)
PASTA_CHANGELOG = RAIZ / "changelog politicas"
HTML_PATH = PASTA_CHANGELOG / "changelog_completo.html"
SAIDA_PATH = PASTA_CHANGELOG / "changelog_politicas_completo.xlsx"

CIAS_ENERGIA = ["COSERN", "CPFL", "ENEL SP", "ENEL RJ", "ENEL CE", "ENEL", "NEO", "RGE"]

# Padrão de regex por CIA — o texto separa "ENEL" do estado ora com espaço, ora com
# hífen ("ENEL SP" / "ENEL - SP"), e "NEO" costuma aparecer colado como "NeoEnergia".
_SEP = r"\s*[-–]?\s*"
CIA_REGEX = {
    "COSERN": re.compile(r"\bCOSERN\b", re.I),
    "CPFL": re.compile(r"\bCPFL\b", re.I),
    "RGE": re.compile(r"\bRGE\b", re.I),
    "ENEL SP": re.compile(rf"\bENEL{_SEP}SP\b", re.I),
    "ENEL RJ": re.compile(rf"\bENEL{_SEP}RJ\b", re.I),
    "ENEL CE": re.compile(rf"\bENEL{_SEP}CE\b", re.I),
    "ENEL": re.compile(rf"\bENEL\b(?!{_SEP}(?:SP|RJ|CE))", re.I),
    "NEO": re.compile(r"\bNEO(?:ENERGIA)?\b", re.I),
}

PRODUTOS_ENERGIA = {"CP Energia", "CDC Energia", "ECP"}

MAPA_PRODUTO = {
    "energia": "CP Energia",
    "cdc_energia": "CDC Energia",
    "cdc energia": "CDC Energia",
    "ecp": "ECP",
    "cp auto": "CP Auto",
    "boleto": "CP Boleto",
    "cp boleto": "CP Boleto",
    "cp refin": "CP Refin",
    "refin": "CP Refin",
    "debito": "Débito",
    "financiamento crefaz": "Financiamento Crefaz",
    "saude": "Saúde",
    "todos": "Todos",
}


def _mapear_ou_bruto(candidato: str) -> str:
    return MAPA_PRODUTO.get(normalizar(candidato), candidato)


def normalizar(texto: str) -> str:
    sem_acento = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return sem_acento.strip().lower()


def extrair_produto(nome_regra: str) -> str:
    # "(Oferta de Produtos) - Oferta X" -> X é o produto real (o resto é o motor genérico)
    m = re.search(r"\(Oferta de Produtos\)\s*[-–]\s*Oferta\s+(.+)$", nome_regra, re.I)
    if m:
        candidato = m.group(1).strip().rstrip(")").strip()
        if " - " in candidato:
            candidato = candidato.split(" - ")[0].strip()
        return _mapear_ou_bruto(candidato)

    # Qualquer "(Oferta X)" ou parêntese "bare" tipo "(ECP)", "(CP AUTO)", "(Todos)"
    # cujo conteúdo bata com um produto conhecido — em qualquer posição da string,
    # pois em "Especificações de Pricing (Energia)" o produto vem no final.
    for grupo in re.findall(r"\(([^)]+)\)", nome_regra):
        candidato = grupo.strip()
        if candidato.lower().startswith("oferta "):
            candidato = candidato[len("oferta "):].strip()
        if " - " in candidato:
            candidato = candidato.split(" - ")[0].strip()
        if normalizar(candidato) in MAPA_PRODUTO:
            return MAPA_PRODUTO[normalizar(candidato)]

    m = re.match(r"Produto[s]?\s+(.+)", nome_regra.strip(), re.I)
    if m:
        return _mapear_ou_bruto(m.group(1).strip())

    if re.search(r"todos\s+os\s+produtos", nome_regra, re.I):
        return "Todos"

    # Nada bateu com o vocabulário conhecido: melhor devolver o conteúdo bruto do
    # primeiro parêntese do que deixar em branco (fácil de filtrar/revisar depois) —
    # exceto se esse conteúdo for só nome(s) de CIA (ex.: "(CPFL/RGE)"), que não é
    # produto e ficaria enganoso; nesse caso é melhor deixar em branco mesmo.
    grupos = re.findall(r"\(([^)]+)\)", nome_regra)
    if grupos:
        candidato = grupos[0].strip()
        if candidato.lower().startswith("oferta "):
            candidato = candidato[len("oferta "):].strip()
        if " - " in candidato:
            candidato = candidato.split(" - ")[0].strip()
        tokens = re.split(r"\s*/\s*|\s*,\s*|\s+e\s+", candidato)
        if all(normalizar(t.strip()) in {normalizar(c) for c in CIAS_ENERGIA} for t in tokens if t.strip()):
            return ""
        return candidato
    return ""


def extrair_cia(produto: str, *textos: str) -> tuple[str, str]:
    if produto not in PRODUTOS_ENERGIA:
        return "", ""
    achadas = []
    junto = " ".join(textos)
    for cia in CIAS_ENERGIA:
        if CIA_REGEX[cia].search(junto) and cia not in achadas:
            achadas.append(cia)
    if not achadas:
        return "Todas", "Todas CIAs"
    if len(achadas) == 1:
        return achadas[0], "CIA específica"
    return ", ".join(achadas), "Múltiplas CIAs"


def parse_versao_header(h2_texto: str) -> tuple[str, str]:
    m = re.match(r"^(.*?)\s*-\s*(\d{2}/\d{2}/\d{4})\s*$", h2_texto.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return h2_texto.strip(), ""


def coletar_entradas_secao(h3_tag) -> list[dict]:
    """A partir de um <h3> (cabeçalho de Política), varre os <li> com <b> (nome da
    regra) pareados com a <ul> de descrição que os segue (como irmã ou aninhada),
    e cai para 'sem nome de regra' quando a seção é só texto corrido."""
    entradas = []
    fim_marcado = {"h2", "h3", "hr"}

    blocos = []
    sib = h3_tag.find_next_sibling()
    while sib is not None and sib.name not in fim_marcado:
        blocos.append(sib)
        sib = sib.find_next_sibling()

    algum_com_b = False
    for bloco in blocos:
        if bloco.name != "ul":
            continue
        for li in bloco.find_all("li", recursive=False):
            b_tag = li.find("b")
            if b_tag is None:
                continue
            algum_com_b = True
            nome_regra = b_tag.get_text(" ", strip=True)

            ul_aninhada = li.find("ul")
            if ul_aninhada is not None:
                texto = ul_aninhada.get_text(" ", strip=True)
            else:
                prox = li.find_next_sibling()
                texto = prox.get_text(" ", strip=True) if prox is not None and prox.name == "ul" else ""

            entradas.append({"nome_regra": nome_regra, "texto": texto})

    if not algum_com_b:
        for bloco in blocos:
            if bloco.name != "ul":
                continue
            for li in bloco.find_all("li", recursive=False):
                texto = li.get_text(" ", strip=True)
                if texto:
                    entradas.append({"nome_regra": "", "texto": texto})

    return entradas


def chave_classificacao(linha: dict) -> str:
    """Chave estável (Versao|Nome da Regra|Texto) usada para persistir a classificação de
    Tipo de Mudança entre execuções — muda se o texto da entrada mudar (evita aplicar uma
    classificação antiga a um conteúdo editado)."""
    base = f"{linha['Versao']}|{linha['Nome da Regra']}|{linha['Texto da Mudanca']}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()[:12]


def parse_changelog(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    linhas = []

    for h2 in soup.find_all("h2"):
        if h2.get_text(strip=True) == "Sumário":
            continue
        versao, data = parse_versao_header(h2.get_text(" ", strip=True))

        sib = h2.find_next_sibling()
        while sib is not None and sib.name != "h2":
            if sib.name == "h3":
                politica = sib.get_text(" ", strip=True)
                if re.search(r"altera[cç][aã]o(?:es)?\s+crivo", politica, re.I):
                    sib = sib.find_next_sibling()
                    continue
                for entrada in coletar_entradas_secao(sib):
                    produto = extrair_produto(entrada["nome_regra"])
                    cia, abrangencia = extrair_cia(produto, entrada["nome_regra"], entrada["texto"])
                    linhas.append(
                        {
                            "Versao": versao,
                            "Data": data,
                            "Produto": produto,
                            "Politica": politica,
                            "Nome da Regra": entrada["nome_regra"],
                            "CIA": cia,
                            "Abrangencia": abrangencia,
                            "Tipo de Mudanca": "",
                            "Texto da Mudanca": entrada["texto"],
                        }
                    )
            sib = sib.find_next_sibling()

    return linhas


def montar_planilha_detalhe(wb: openpyxl.Workbook, linhas: list[dict]) -> None:
    ws = wb.active
    ws.title = "Changelog Completo"
    colunas = [
        "Versao", "Data", "Produto", "Politica", "Nome da Regra",
        "CIA", "Abrangencia", "Tipo de Mudanca", "Texto da Mudanca",
    ]
    cabecalho = [
        "Versão", "Data", "Produto", "Política", "Nome da Regra",
        "CIA", "Abrangência", "Tipo de Mudança", "Texto da Mudança",
    ]
    ws.append(cabecalho)
    for linha in linhas:
        ws.append([linha[c] for c in colunas])

    n_colunas = len(cabecalho)
    for col in range(1, n_colunas + 1):
        c = ws.cell(row=1, column=col)
        c.fill = COR_CABECALHO
        c.font = FONTE_CABECALHO
        c.alignment = Alignment(horizontal="center")

    idx_produto, idx_cia, idx_tipo = 3, 6, 8
    for i, linha in enumerate(linhas):
        r = i + 2
        faixa = COR_FAIXA_PAR if r % 2 == 0 else COR_FAIXA_IMPAR
        for col in range(1, n_colunas + 1):
            cel = ws.cell(row=r, column=col)
            cel.border = BORDA_FINA
            if col == idx_produto and linha["Produto"] in PRODUTO_CORES:
                cel.fill = _fill(PRODUTO_CORES[linha["Produto"]])
            elif col == idx_cia and _cor_cia(linha["CIA"]):
                cel.fill = _fill(_cor_cia(linha["CIA"]))
            elif col == idx_tipo and linha["Tipo de Mudanca"] in TIPO_CORES:
                cel.fill = _fill(TIPO_CORES[linha["Tipo de Mudanca"]])
            else:
                cel.fill = faixa

    larguras = [24, 13, 22, 30, 45, 16, 16, 18, 70]
    for idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_colunas)}{ws.max_row}"


def montar_planilha_resumo(wb: openpyxl.Workbook, linhas: list[dict]) -> None:
    ws = wb.create_sheet("Resumo")
    fonte_titulo = Font(name="Calibri", size=12, bold=True, color="1F3864")
    fonte_subcabecalho = Font(name="Calibri", size=11, bold=True)

    def bloco(titulo: str, contagem: Counter, linha_inicial: int) -> int:
        ws.cell(row=linha_inicial, column=1, value=titulo).font = fonte_titulo
        c_item = ws.cell(row=linha_inicial + 1, column=1, value="Item")
        c_qtd = ws.cell(row=linha_inicial + 1, column=2, value="Qtd")
        c_item.font = c_qtd.font = fonte_subcabecalho
        r = linha_inicial + 2
        for item, qtd in contagem.most_common():
            if not item:
                continue
            ws.cell(row=r, column=1, value=item)
            ws.cell(row=r, column=2, value=qtd)
            r += 1
        return r + 1

    por_produto = Counter(l["Produto"] for l in linhas)
    por_ano = Counter(l["Data"][-4:] for l in linhas if l["Data"])
    por_cia = Counter(cia for l in linhas if l["CIA"] for cia in l["CIA"].split(", "))
    por_abrangencia = Counter(l["Abrangencia"] for l in linhas if l["Abrangencia"])
    por_tipo = Counter(l["Tipo de Mudanca"] for l in linhas if l["Tipo de Mudanca"])

    r = 1
    r = bloco("Por Produto", por_produto, r)
    r = bloco("Por Ano", por_ano, r)
    r = bloco("Por CIA (apenas produtos de Energia)", por_cia, r)
    r = bloco("Por Abrangência (apenas produtos de Energia)", por_abrangencia, r)
    if por_tipo:
        bloco("Por Tipo de Mudança", por_tipo, r)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 10


def montar_planilha_produto_cia(wb: openpyxl.Workbook, linhas: list[dict]) -> None:
    ws = wb.create_sheet("Produto x CIA")
    matriz: dict[str, Counter] = {}
    cias_vistas: set[str] = set()
    for l in linhas:
        if not l["CIA"]:
            continue
        produto = l["Produto"]
        # uma linha "Múltiplas CIAs" conta para cada CIA individual que ela cita
        # (não vira uma coluna própria "A, B"), igual ao changelog_energia.xlsx original.
        for cia in l["CIA"].split(", "):
            matriz.setdefault(produto, Counter())[cia] += 1
            cias_vistas.add(cia)

    cias_ordenadas = sorted(cias_vistas)
    fonte_cabecalho = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=1, column=1, value="Produto \\ CIA").font = fonte_cabecalho
    for c, cia in enumerate(cias_ordenadas, start=2):
        ws.cell(row=1, column=c, value=cia).font = fonte_cabecalho
    for r, produto in enumerate(sorted(matriz), start=2):
        ws.cell(row=r, column=1, value=produto)
        for c, cia in enumerate(cias_ordenadas, start=2):
            qtd = matriz[produto].get(cia)
            if qtd:
                ws.cell(row=r, column=c, value=qtd)

    ws.column_dimensions["A"].width = 26
    for c in range(2, 2 + len(cias_ordenadas)):
        ws.column_dimensions[get_column_letter(c)].width = 12


def main() -> None:
    html = HTML_PATH.read_text(encoding="utf-8")
    linhas = parse_changelog(html)

    wb = openpyxl.Workbook()
    montar_planilha_detalhe(wb, linhas)
    montar_planilha_resumo(wb, linhas)
    montar_planilha_produto_cia(wb, linhas)

    SAIDA_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SAIDA_PATH)

    sem_produto = sum(1 for l in linhas if not l["Produto"])
    print(f"{len(linhas)} linhas geradas -> {SAIDA_PATH}")
    print(f"{sem_produto} linha(s) sem Produto identificado (revisar manualmente).")


if __name__ == "__main__":
    main()
